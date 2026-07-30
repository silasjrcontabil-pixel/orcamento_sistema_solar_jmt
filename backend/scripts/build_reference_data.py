"""Gera dados de referência estáticos usados pelo backend em runtime.

Roda uma única vez (localmente) e commita o resultado em app/data/. Não deve
ser executado em produção — evita a dependência de openpyxl/parsing pesado
no deploy do Render.

Uso: python scripts/build_reference_data.py
"""
import json
import shutil
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
BACKEND = Path(__file__).resolve().parent.parent
DATA_DIR = BACKEND / "app" / "data"

MUNICIPIOS_JSON = ROOT / "municipios.json"
ESTADOS_JSON = ROOT / "estados.json"
XLSX = ROOT / "Calcular Radicao" / "Munipios_Longitude.xlsx"
IRRADIANCIA_CSV = ROOT / "Calcular Radicao" / "Irradiação no Plano Inclinado.csv"


def parse_br_float(value: str) -> float:
    return float(str(value).strip().replace(",", "."))


def build_municipios_geo() -> dict:
    with open(MUNICIPIOS_JSON, encoding="utf-8") as f:
        municipios = json.load(f)  # {cod_ibge: nome}

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    ws = wb["LOCALIDADES"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(header)}

    # cod_ibge (7 dígitos) -> {uf, lat, lon} a partir da planilha
    geo_by_ibge = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod_ibge = row[col["COD_IBGE"]]
        if cod_ibge is None:
            continue
        cod_ibge = str(cod_ibge).strip()
        try:
            lat = parse_br_float(row[col["LATITUDE"]])
            lon = parse_br_float(row[col["LONGITUDE"]])
        except (TypeError, ValueError):
            continue
        uf = row[col["UF"]]
        # planilha tem 1 linha por faixa de CEP: manter a primeira ocorrência por município
        if cod_ibge not in geo_by_ibge:
            geo_by_ibge[cod_ibge] = {"uf": uf, "lat": lat, "lon": lon}

    result = {}
    for cod_ibge, nome in municipios.items():
        geo = geo_by_ibge.get(cod_ibge)
        if geo is None:
            continue
        result[cod_ibge] = {
            "nome": nome,
            "uf": geo["uf"],
            "lat": geo["lat"],
            "lon": geo["lon"],
        }

    missing = len(municipios) - len(result)
    print(f"municipios_geo: {len(result)} municípios resolvidos, {missing} sem match na planilha")
    return result


def main():
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    municipios_geo = build_municipios_geo()
    out_path = DATA_DIR / "municipios_geo.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(municipios_geo, f, ensure_ascii=False)
    print(f"Escrito {out_path} ({out_path.stat().st_size / 1024:.1f} KB)")

    dest_csv = DATA_DIR / "irradiancia.csv"
    shutil.copyfile(IRRADIANCIA_CSV, dest_csv)
    print(f"Copiado {dest_csv} ({dest_csv.stat().st_size / 1024:.1f} KB)")


if __name__ == "__main__":
    main()
